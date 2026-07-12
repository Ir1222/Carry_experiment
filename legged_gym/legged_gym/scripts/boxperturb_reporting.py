"""Shared CSV and terminal reporting for carry-box perturbation evaluation."""

import csv
import json
import math
import os
from collections import defaultdict
from statistics import mean, stdev


FORCE_METRICS = (
    "left_fn_raw_N", "right_fn_raw_N", "left_ft_raw_N", "right_ft_raw_N",
    "left_fn_ema_N", "right_fn_ema_N", "left_ft_ema_N", "right_ft_ema_N",
)


def _finite(rows, key):
    return [
        float(row[key]) for row in rows
        if key in row and math.isfinite(float(row[key]))
    ]


def _stat(rows, key, operation="mean"):
    values = _finite(rows, key)
    if not values:
        return float("nan")
    return max(values) if operation == "max" else sum(values) / len(values)


def summarize_force_trace(trace):
    """Return per-trial pre/pulse/recovery response metrics from raw trace rows."""
    pre = [r for r in trace if r.get("phase") == "pre" and r.get("force_decomposition_valid") == 1][-40:]
    pulse = [r for r in trace if r.get("phase") == "pulse"]
    recovery = [r for r in trace if r.get("phase") == "recovery"]
    response = {}
    for key in FORCE_METRICS:
        pre_mean = _stat(pre, key)
        pulse_mean = _stat(pulse, key)
        response[f"{key}_pre_mean"] = pre_mean
        response[f"{key}_pulse_mean"] = pulse_mean
        response[f"{key}_pulse_peak"] = _stat(pulse, key, "max")
        response[f"{key}_recovery_mean"] = _stat(recovery, key)
        response[f"{key}_pulse_delta_from_pre"] = (
            pulse_mean - pre_mean
            if math.isfinite(pre_mean) and math.isfinite(pulse_mean)
            else float("nan")
        )
    all_response = pulse + recovery
    response.update(
        force_valid_fraction=(
            sum(int(r.get("force_decomposition_valid", 0)) for r in all_response) / len(all_response)
            if all_response else float("nan")
        ),
        pulse_force_valid_fraction=(
            sum(int(r.get("force_decomposition_valid", 0)) for r in pulse) / len(pulse)
            if pulse else float("nan")
        ),
        force_baseline_sample_count=len(pre),
        force_baseline_unavailable=int(len(pre) < 40),
        force_closure_residual_pulse_mean=_stat(pulse, "force_closure_residual"),
        left_rho_pulse_mean=_stat(pulse, "left_rho_raw"),
        right_rho_pulse_mean=_stat(pulse, "right_rho_raw"),
        normal_load_asymmetry_pulse_mean=_stat(pulse, "normal_load_asymmetry"),
        external_impulse_Ns=(max(_finite(trace, "force_impulse_Ns"), default=0.0)),
    )
    if pulse:
        last = pulse[-1]
        for key in (
            "force_uncapped_peak_N", "force_peak_N", "force_cap_used",
            "perturb_direction_world_x", "perturb_direction_world_y", "perturb_direction_world_z",
            "force_point_box_x", "force_point_box_y", "force_point_box_z",
            "force_point_world_x", "force_point_world_y", "force_point_world_z",
            "external_torque_norm_Nm", "pulse_duration_s", "actual_force_scale",
        ):
            response[key] = last.get(key, float("nan"))
    return response


def print_trial_terminal(trial, prefix="BoxPerturb"):
    direction = trial.get("direction", "unknown")
    beta = float(trial.get("requested_beta", float("nan")))
    print(
        f"[{prefix}:Perturb] direction={direction} beta={beta:.3f} "
        f"world=({float(trial.get('perturb_direction_world_x', float('nan'))):.4f},"
        f"{float(trial.get('perturb_direction_world_y', float('nan'))):.4f},"
        f"{float(trial.get('perturb_direction_world_z', float('nan'))):.4f}) "
        f"peak={float(trial.get('peak_force_N', trial.get('force_peak_N', float('nan')))):.4f}N"
    )


def write_csv(path, rows):
    if not rows:
        return
    fields = []
    for row in rows:
        for key in row:
            if key not in fields:
                fields.append(key)
    with open(path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def aggregate_trials(trials):
    groups = defaultdict(list)
    for row in trials:
        groups[(
            row["model"],
            row["direction"],
            row["requested_beta"],
            row.get("force_point_mode", "com"),
            row.get("force_point_label", "com"),
            row.get("pulse_duration_s", 0.10),
            row.get("pulse_profile", "half_sine"),
        )].append(row)
    metrics = (
        "recovery_success", "pulse_hold_retention", "post_confirmed_ratio",
        "force_valid_fraction", "force_closure_residual_pulse_mean",
    ) + tuple(
        f"{key}_{suffix}"
        for key in FORCE_METRICS
        for suffix in ("pre_mean", "pulse_mean", "pulse_peak", "pulse_delta_from_pre")
    )
    result = []
    for (model, direction, beta, point_mode, point_label, duration, profile), rows in sorted(groups.items()):
        item = {
            "model": model, "direction": direction, "beta": beta,
            "force_point_mode": point_mode, "force_point_label": point_label,
            "pulse_duration_s": duration, "pulse_profile": profile,
            "trials": len(rows),
        }
        item["precondition_success_rate"] = sum(int(r.get("precondition_success", 0)) for r in rows) / len(rows)
        for metric in metrics:
            values = _finite(rows, metric)
            item[f"{metric}_mean"] = mean(values) if values else float("nan")
            item[f"{metric}_std"] = stdev(values) if len(values) > 1 else 0.0
        result.append(item)
    return result


def build_boundary_rows(trials):
    groups = defaultdict(list)
    for row in trials:
        groups[(
            row["model"],
            row["direction"],
            row.get("force_point_mode", "com"),
            row.get("force_point_label", "com"),
            row.get("pulse_duration_s", 0.10),
            row.get("pulse_profile", "half_sine"),
        )].append(row)
    rows = []
    for (model, direction, point_mode, point_label, duration, profile), items in sorted(groups.items()):
        valid = [r for r in items if int(r.get("precondition_success", 0)) == 1]
        passed = []
        failed = []
        degraded = []
        for r in valid:
            beta = float(r["requested_beta"])
            primary_failure = bool(int(r.get("termination", 0))) or (
                math.isfinite(float(r.get("post_confirmed_ratio", float("nan"))))
                and float(r.get("post_confirmed_ratio", 0.0)) < 0.5
            )
            degradation_failure = (
                int(r.get("recovery_success", 0)) == 0
                or float(r.get("pulse_force_valid_fraction", 1.0)) < 0.5
                or float(r.get("hand_box_rel_speed_peak_mps", 0.0)) > 1.0
            )
            if primary_failure:
                failed.append(beta)
            else:
                passed.append(beta)
            if degradation_failure:
                degraded.append(beta)
        rows.append({
            "model": model,
            "direction": direction,
            "force_point_mode": point_mode,
            "force_point_label": point_label,
            "pulse_duration_s": duration,
            "pulse_profile": profile,
            "trials": len(items),
            "precondition_success_rate": (
                sum(int(r.get("precondition_success", 0)) for r in items) / len(items)
                if items else float("nan")
            ),
            "max_pass_beta": max(passed) if passed else float("nan"),
            "min_primary_failure_beta": min(failed) if failed else float("nan"),
            "min_degradation_beta": min(degraded) if degraded else float("nan"),
            "valid_trials": len(valid),
        })
    return rows


def write_variable_dictionary(path):
    content = """# Box perturbation variable dictionary

All force values are SI units unless stated otherwise.

| Variable | Meaning | Unit | Directionality |
|---|---|---:|---|
| `direction` | Nominal perturbation direction token. `box_x/y` are box-local axes recomputed from the current box orientation every physics substep; `world_z` is gravity-aligned. | - | diagnostic |
| `perturb_direction_local_{x,y,z}` | Stored local/world semantic direction. For `box_x/y` this is box-local; for `world_z` this is already world Z. | - | diagnostic |
| `perturb_direction_is_world` | `1` for `±world_z`, `0` for box-local directions. | 0/1 | diagnostic |
| `requested_beta` | Requested force scale. `F_uncapped = beta * box_mass_kg * 9.81`. | - | larger is stronger |
| `force_uncapped_peak_N` | Peak force before cap. | N | larger is stronger |
| `force_peak_N` | Actual capped peak force. | N | larger is stronger |
| `force_peak_cap_N` | Peak cap; NaN means no cap. | N | constraint |
| `force_point_mode` | `com`, `box_surface_grid`, or `box_surface_random`. | - | diagnostic |
| `force_point_label` | Grid point label such as `face_center`, `face_upper`, `face_left_edge`. | - | diagnostic |
| `force_point_box_{x,y,z}` | Application point offset in box local frame. | m | diagnostic |
| `force_point_world_{x,y,z}` | Application point in world/env frame. | m | diagnostic |
| `moment_arm_world_{x,y,z}` | `force_point_world - box_com_world`. | m | diagnostic |
| `external_torque_world_Nm_{x,y,z}` | Torque induced by off-center force, `tau = r × F`. | N m | larger means stronger rotational disturbance |
| `pulse_duration_s` | Force pulse duration. | s | longer means larger impulse for same peak/profile |
| `pulse_profile` | `half_sine`, `ramp_hold`, `multi_pulse`, or `jittered_half_sine`. | - | diagnostic |
| `actual_force_scale` | Per-substep profile multiplier applied to peak force. | - | diagnostic |
| `force_impulse_Ns` | Cumulative integral of external force norm. | N s | larger is stronger |
| `left/right_fn_raw_N` | Hand net contact force projected onto estimated box face normal. | N | diagnostic |
| `left/right_ft_raw_N` | Tangential component magnitude after normal projection. | N | diagnostic |
| `left/right_rho_raw` | `Ft / (Fn + eps)`. Can explode when Fn≈0; use only with valid samples. | - | diagnostic |
| `force_decomposition_valid` | Validity gate for Fn/Ft projection. | 0/1 | higher is more reliable |
| `force_closure_residual` | Closure audit between hand net force and box net contact force. | - | lower is better |
| `post_confirmed_ratio` | Confirmed-carry fraction after recovery window. | - | higher is better |
| `recovery_success` | Whether confirmed carry recovered for the required streak. | 0/1 | higher is better |
| `termination` | Whether the episode terminated after perturbation. | 0/1 | lower is better |
| `max_pass_beta` | Largest tested beta without primary failure in a boundary group. | - | higher is better |
| `min_primary_failure_beta` | Smallest beta causing termination or `post_confirmed_ratio < 0.5`. | - | higher is better |
| `min_degradation_beta` | Smallest beta causing recovery/contact/relative-speed degradation. | - | higher is better |

Important caveat: `Fn/Ft` is a projection of each hand rigid-body net contact force on an estimated locked box-face normal. It is not strict pairwise hand-box contact force.
"""
    with open(path, "w", encoding="utf-8") as handle:
        handle.write(content)


def write_analysis_workbook(output_dir, trials, summary, boundary):
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference
    except Exception:
        return False
    workbook = Workbook()
    for sheet in list(workbook.worksheets):
        workbook.remove(sheet)

    def add_sheet(name, rows):
        ws = workbook.create_sheet(name[:31])
        if not rows:
            return ws
        fields = []
        for row in rows:
            for key in row:
                if key not in fields:
                    fields.append(key)
        ws.append(fields)
        for row in rows:
            values = []
            for key in fields:
                value = row.get(key, "")
                if isinstance(value, float) and not math.isfinite(value):
                    value = ""
                values.append(value)
            ws.append(values)
        return ws

    ws_summary = add_sheet("summary", summary)
    add_sheet("trials", trials)
    add_sheet("boundary", boundary)
    if ws_summary.max_row > 1 and ws_summary.max_column > 1:
        chart = BarChart()
        chart.title = "Precondition success by cell"
        headers = [cell.value for cell in ws_summary[1]]
        if "precondition_success_rate" in headers:
            col = headers.index("precondition_success_rate") + 1
            data = Reference(ws_summary, min_col=col, min_row=1, max_row=ws_summary.max_row)
            chart.add_data(data, titles_from_data=True)
            ws_summary.add_chart(chart, "J2")
    workbook.save(os.path.join(output_dir, "analysis.xlsx"))
    return True


def write_run_files(
    output_dir, trials, traces, metadata, summary=None, write_force_trace=True
):
    os.makedirs(output_dir, exist_ok=True)
    summary = summary or aggregate_trials(trials)
    boundary = build_boundary_rows(trials)
    if write_force_trace:
        write_csv(os.path.join(output_dir, "force_trace.csv"), traces)
    write_csv(os.path.join(output_dir, "trials.csv"), trials)
    write_csv(os.path.join(output_dir, "summary.csv"), summary)
    write_csv(os.path.join(output_dir, "boundary.csv"), boundary)
    write_variable_dictionary(os.path.join(output_dir, "variable_dictionary.md"))
    write_analysis_workbook(output_dir, trials, summary, boundary)
    with open(os.path.join(output_dir, "run_metadata.json"), "w", encoding="utf-8") as handle:
        json.dump(metadata, handle, indent=2, ensure_ascii=False, allow_nan=True)
